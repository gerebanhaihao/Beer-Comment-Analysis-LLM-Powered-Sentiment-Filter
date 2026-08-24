"""Styled Excel output."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from beer_sentiment.config import AppConfig
from beer_sentiment.models import Label


def _excel_safe(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", str(value))


def write_colored_excel(
    path: str | Path,
    rows: list[dict[str, Any]],
    labels: list[Label],
    headers: list[str],
    config: AppConfig,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "舆情"
    worksheet.append(headers)

    font_name = config.excel.get("font", "宋体")
    font_size = int(config.excel.get("font_size", 11))
    row_height = int(config.excel.get("row_height", 14))
    column_width = float(config.excel.get("column_width", 8.42))
    font = Font(name=font_name, size=font_size)
    alignment = Alignment(vertical="top")
    blue_fill = PatternFill(fill_type="solid", fgColor=config.colors["blue"])
    yellow_fill = PatternFill(fill_type="solid", fgColor=config.colors["yellow"])

    for cell in worksheet[1]:
        cell.font = font
        cell.alignment = alignment
    worksheet.row_dimensions[1].height = row_height

    for row, label in zip(rows, labels):
        worksheet.append([_excel_safe(row.get(header, "")) for header in headers])
        row_number = worksheet.max_row
        worksheet.row_dimensions[row_number].height = row_height
        for cell in worksheet[row_number]:
            cell.font = font
            cell.alignment = alignment
        if label == Label.BLUE:
            for cell in worksheet[row_number]:
                cell.fill = blue_fill
        elif label == Label.YELLOW:
            for cell in worksheet[row_number]:
                cell.fill = yellow_fill

    for col_index, _ in enumerate(headers, start=1):
        worksheet.column_dimensions[get_column_letter(col_index)].width = column_width
    worksheet.freeze_panes = "A2"
    workbook.save(path)
