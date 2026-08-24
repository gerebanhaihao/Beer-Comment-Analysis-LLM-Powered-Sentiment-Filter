import csv
from pathlib import Path

from openpyxl import load_workbook

from beer_sentiment.pipeline.run import run_file


def _write_csv(path: Path, rows: list[list[str]]) -> None:
    headers = ["数据范围", "发帖时间", "标题", "正文", "封面OCR", "内容OCR", "情感"]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def test_run_file_colors_excel(tmp_path, config, mock_judge):
    source = tmp_path / "quark__2026-08-24 091500.csv"
    rows = [
        ["百威（本品）-产品", "2026-08-24 09:20:00", "百威避雷", "昨晚喝的百威太难喝了，一股水味。", "", "", "负面"],
        ["青岛（竞品）-产品", "2026-08-24 08:50:00", "青岛翻车", "青岛啤酒抽检不合格被罚款，大家注意。", "", "", "负面"],
        ["百威（本品）-产品", "2026-08-23 19:00:00", "周末愉快", "朋友聚会喝科罗娜，气氛很好。", "", "", "正面"],
        ["啤酒行业", "2026-08-24 09:10:00", "广告", "精酿店里老板宣传自家原产地酿造，说工业勾兑酒别喝。", "", "", "中性"],
        ["雪花（竞品）-产品", "2026-08-24 06:40:00", "行业观察", "雪花卖不动了，库存压力很大。", "", "", "负面"],
        ["百威（本品）-产品", "2026-08-24 08:15:00", "新品发布", "百威发布全新无醇系列，主打低卡健康。", "", "", "正面"],
    ]
    _write_csv(source, rows)

    summary = run_file(source, tmp_path / "out", "morning", "2026-08-24", mock_judge, config)
    assert summary.total_rows == 6
    assert summary.blue_rows == 1
    assert summary.yellow_rows == 2
    assert len(summary.low_confidence_rows) == 0

    out = Path(summary.output_path)
    assert out.name == "品牌上午__2026-08-24 091500.xlsx"
    workbook = load_workbook(out)
    worksheet = workbook.active
    fills = [
        worksheet.cell(row=row, column=1).fill.start_color.rgb
        for row in range(2, worksheet.max_row + 1)
    ]
    assert fills[0] == "FF00B0F0"
    assert fills[1] == "FFFFFF00"
    assert fills[2] == "00000000"
    assert fills[3] == "00000000"
    assert fills[4] == "FFFFFF00"
    assert fills[5] == "00000000"


def test_run_file_flags_low_confidence(tmp_path, config, mock_judge):
    source = tmp_path / "quark__2026-08-24 090000.csv"
    rows = [
        ["啤酒行业", "2026-08-24 09:40:00", "求助", "现在的啤酒都是勾兑的吗？有没有懂行的说说。", "", "", "中性"],
    ]
    _write_csv(source, rows)

    summary = run_file(source, tmp_path / "out", "morning", "2026-08-24", mock_judge, config)
    assert len(summary.low_confidence_rows) == 1
    assert summary.yellow_rows == 0
